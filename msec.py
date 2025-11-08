
import os
import sys
import requests
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


TIPOS_TITULOS = ["ltn", "ntn-c", "lft", "ntn-b", "ntn-f"]
URL_BASE = "https://www.anbima.com.br/informacoes/merc-sec/resultados/msec_{data_url}_{tipo}.asp"

# Mapeamento de meses em português
MESES_PT = {
    1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun',
    7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'
}

# Nomes das abas para a planilha
NOMES_ABAS = {
    "ltn": "LTN",
    "ntn-c": "NTN-C",
    "lft": "LFT",
    "ntn-b": "NTN-B",
    "ntn-f": "NTN-F"
}

# A3 na planilha
TIPO_PAPEL = {
    "ltn": "Papel PREFIXADO",
    "ntn-c": "Papel IGP-M",
    "lft": "Papel POS-SELIC",
    "ntn-b": "Papel IPCA",
    "ntn-f": "Papel PREFIXADO"
}

# C3 na planilha
NOME_PAPEL = {
    "ltn": "LTN - Taxa (% a.a.)/252",
    "ntn-c": "NTN-C - Taxa (% a.a.)/252",
    "lft": "LFT - Rentabilidade (% a.a.)/252",
    "ntn-b": "NTN-B - Taxa (% a.a.)/252",
    "ntn-f": "NTN-F - Taxa (% a.a.)/252"
}


def formatar_data_anbima(data_obj: datetime) -> str:
    """Formata data no padrão ANBIMA: 31out2025"""
    dia = data_obj.strftime('%d')
    mes = MESES_PT[data_obj.month]
    ano = data_obj.strftime('%Y')
    return f"{dia}{mes}{ano}"


def contar_dias_uteis_entre_datas(data_inicio: datetime, data_fim: datetime) -> int:
    """Conta o número de dias úteis entre duas datas (excluindo fins de semana)"""
    dias_uteis = 0
    data_atual = data_inicio
    
    while data_atual < data_fim:
        # Se não é sábado (5) nem domingo (6)
        if data_atual.weekday() < 5:
            dias_uteis += 1
        data_atual += timedelta(days=1)
    
    return dias_uteis


def calcular_dia_util_anterior() -> tuple[str, str]:
    """Calcula o dia útil anterior (D-X)"""
    hoje = datetime.now()
    delta_num = 1
    
    if hoje.weekday() == 0:
        delta_num = 3
    elif hoje.weekday() == 6:
        delta_num = 2
        
    data_anterior = hoje - timedelta(days=delta_num)
    data_str = data_anterior.strftime('%d/%m/%Y')
    delta_str = f"D-{delta_num}"
    
    return data_str, delta_str


def calcular_data_minima_permitida() -> str:
    """Calcula a data mínima permitida (5 dias úteis atrás)"""
    hoje = datetime.now()
    dias_uteis_contados = 0
    data_atual = hoje
    
    while dias_uteis_contados < 5:
        data_atual -= timedelta(days=1)
        # Se não é sábado (5) nem domingo (6)
        if data_atual.weekday() < 5:
            dias_uteis_contados += 1
    
    return data_atual.strftime('%d/%m/%Y')


def calcular_data_consulta() -> str:
    """Determina a data de consulta com validações"""
    DATA_HOJE = datetime.now().date()
    data_input = sys.argv[1] if len(sys.argv) > 1 else None
    
    if data_input:
        try:
            data_obj = datetime.strptime(data_input, '%d/%m/%Y').date()
            
            if data_obj >= DATA_HOJE:
                print(f"❌ Data precisa ser anterior a {DATA_HOJE.strftime('%d/%m/%Y')}.")
                sys.exit(1)
            
            dia_da_semana = data_obj.weekday()
            if dia_da_semana == 5:
                print(f"❌ {data_input} foi um SÁBADO. Selecione um dia útil.")
                sys.exit(1)
            elif dia_da_semana == 6:
                print(f"❌ {data_input} foi um DOMINGO. Selecione um dia útil.")
                sys.exit(1)
            
            # Verificar se está dentro dos últimos 5 dias úteis
            dias_uteis_atras = contar_dias_uteis_entre_datas(
                datetime.combine(data_obj, datetime.min.time()),
                datetime.combine(DATA_HOJE, datetime.min.time())
            )
            
            if dias_uteis_atras > 5:
                data_minima = calcular_data_minima_permitida()
                print(f"❌ Histórico da ANBIMA é de apenas 5 dias úteis.")
                print(f"   Selecione uma data posterior a {data_minima}")
                sys.exit(1)
            
            return data_input
            
        except ValueError:
            print(f"❌ Utilize o padrão dd/mm/aaaa.")
            sys.exit(1)
    else:
        data_calculada, delta = calcular_dia_util_anterior()
        print(f"🗓️ Usando último dia útil ({delta}): {data_calculada}")
        return data_calculada


def gerar_link_anbima(data_consulta: str, tipo_titulo: str) -> str:
    """Constrói a URL da ANBIMA"""
    data_obj = datetime.strptime(data_consulta, '%d/%m/%Y')
    data_url = formatar_data_anbima(data_obj)
    return URL_BASE.format(data_url=data_url, tipo=tipo_titulo)


def extrair_dados_tabela(html_content: str) -> list:
    """Extrai os dados numéricos da tabela HTML da ANBIMA"""
    soup = BeautifulSoup(html_content, 'html.parser')
    tabela = soup.find('table', {'border': True})
    
    if not tabela:
        return []
    
    dados = []
    for tr in tabela.find_all('tr'):
        celulas = tr.find_all('td')
        if not celulas:
            continue
        
        linha = [td.get_text(strip=True) for td in celulas]
        
        if linha and linha[0].isdigit():
            dados.append(linha)
    
    return dados


def converter_valor_celula(valor: str, coluna: int):
    """Converte o valor conforme o tipo de coluna"""
    if coluna == 1:
        try:
            return int(valor)
        except:
            return valor
    
    elif coluna in [2, 3]:
        return valor
    
    elif coluna == 7:
        if '.' in valor and ',' in valor:
            return valor
        else:
            try:
                return float(valor.replace(',', '.'))
            except:
                return valor
    
    else:
        try:
            return float(valor.replace(',', '.'))
        except:
            return valor


def aplicar_formatacao_linha_dados(ws, linha_idx, max_col=11):
    """Aplica formatação padrão para linhas de dados"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(linha_idx, col_idx)
        
        # Bordas
        cell.border = thin_border
        
        # Alinhamento centralizado
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formato de número baseado na coluna
        if col_idx == 1:  # Código SELIC
            cell.number_format = '0'
        elif col_idx in [2, 3]:  # Datas
            cell.number_format = '@'  # Texto
        elif col_idx == 7:  # PU
            # Verifica se é string (valor grande) ou número
            if isinstance(cell.value, str):
                cell.number_format = '@'
            else:
                cell.number_format = '#,##0.000000'
        else:  # Taxas
            cell.number_format = '0.0000'


def criar_arquivo_excel(data_consulta: str, dados_titulos: dict):
    """Cria o arquivo Excel com os dados extraídos"""
    # Nome do arquivo
    data_obj = datetime.strptime(data_consulta, '%d/%m/%Y')
    data_arquivo = formatar_data_anbima(data_obj)
    
    # Criar pasta 'relatorios' se não existir
    pasta_relatorio = 'relatorios'
    if not os.path.exists(pasta_relatorio):
        os.makedirs(pasta_relatorio)
        print(f"📁 Pasta '{pasta_relatorio}' criada.")
    
    # Caminho completo do arquivo
    nome_arquivo = os.path.join(pasta_relatorio, f"msec_{data_arquivo}.xlsx")
    
    # Carregar modelo
    modelo_path = 'modelo.xlsx'
    wb = openpyxl.load_workbook(modelo_path)
    ws_modelo = wb['modelo']
    
    # Contador
    abas_criadas = 0
    total_registros = 0
    
    # Criar uma aba para cada tipo de título
    for tipo in TIPOS_TITULOS:
        dados = dados_titulos.get(tipo, [])
        
        if not dados:
            print(f"  ⚠️ Sem dados para {tipo.upper()}")
            continue
        
        nome_aba = NOMES_ABAS[tipo]
        ws = wb.create_sheet(title=nome_aba)
        
        # Copiar todas as células do modelo
        for row_idx, row in enumerate(ws_modelo.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                new_cell = ws.cell(row_idx, col_idx)
                
                # Copiar valor
                new_cell.value = cell.value
                
                # Copiar formatação
                if cell.font:
                    new_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                
                if cell.border:
                    new_cell.border = Border(
                        left=Side(style=cell.border.left.style) if cell.border.left else None,
                        right=Side(style=cell.border.right.style) if cell.border.right else None,
                        top=Side(style=cell.border.top.style) if cell.border.top else None,
                        bottom=Side(style=cell.border.bottom.style) if cell.border.bottom else None
                    )
                
                if cell.fill:
                    new_cell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
                
                if cell.alignment:
                    new_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text
                    )
                
                if cell.number_format:
                    new_cell.number_format = cell.number_format
        
        # Copiar células mescladas
        for merged_cell_range in ws_modelo.merged_cells.ranges:
            ws.merge_cells(str(merged_cell_range))
        
        # Aplicar largura de 15 a todas as colunas
        for col_idx in range(1, 12):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # Ajustar altura da linha 5
        ws.row_dimensions[4].height = 30
        
        # Copiar alturas das demais linhas do modelo
        for row in ws_modelo.row_dimensions:
            if row != 5 and ws_modelo.row_dimensions[row].height:
                ws.row_dimensions[row].height = ws_modelo.row_dimensions[row].height
        
        # Adicionar tipo de papel na célula A3 (linha 3, coluna 1)
        ws.cell(3, 1).value = TIPO_PAPEL.get(tipo, "")
        
        # Atualizar nome do papel na linha 3, coluna C
        ws.cell(3, 3).value = NOME_PAPEL.get(tipo, "")
        
        # Adicionar data na célula K1 (formato: 06nov2025) e centralizar
        cell_k1 = ws.cell(1, 11)
        cell_k1.value = data_arquivo
        cell_k1.alignment = Alignment(horizontal='center', vertical='center')
        
        # Inserir dados a partir da linha 6
        linha_inicio = 6
        for idx, linha_dados in enumerate(dados):
            linha_atual = linha_inicio + idx
            
            for col_idx, valor in enumerate(linha_dados, start=1):
                if col_idx > 11:
                    break
                
                cell = ws.cell(linha_atual, col_idx)
                cell.value = converter_valor_celula(valor, col_idx)
            
            # Aplicar formatação na linha
            aplicar_formatacao_linha_dados(ws, linha_atual)
        
        print(f"  ✓ {nome_aba:5}")
        abas_criadas += 1
        total_registros += len(dados)
    
    # Remover aba modelo
    if 'modelo' in wb.sheetnames:
        wb.remove(wb['modelo'])
    
    if abas_criadas == 0:
        print("❌ Nenhuma aba foi criada. Verifique os dados.")
        return None
    
    # Salvar arquivo
    wb.save(nome_arquivo)
    print(f"✅ Arquivo salvo: {nome_arquivo}")
    return nome_arquivo


def processar_titulos(data_consulta: str):
    """Processa todos os tipos de títulos"""
    dados_titulos = {}
    
    print(f"📊 Extraindo dados de {data_consulta}...")
    
    for titulo in TIPOS_TITULOS:
        url = gerar_link_anbima(data_consulta, titulo)
        
        try:
            response = requests.get(url, timeout=15)
            response.raise_for_status()
            
            dados = extrair_dados_tabela(response.text)
            dados_titulos[titulo] = dados
                
        except Exception as e:
            print(f"  ✗ {titulo.upper()}: Erro ao extrair dados")
            dados_titulos[titulo] = []
    
    return dados_titulos


# --- Bloco Principal ---
if __name__ == "__main__":
    print("="*70)
    print(" WEBSCRAPING ANBIMA - MERCADO SECUNDÁRIO DE TÍTULOS PÚBLICOS")
    print("   Criado por: Jader Santos (msantos.jader@gmail.com)")
    print("="*70)
    
    DATA_CONSULTA = calcular_data_consulta()
    dados = processar_titulos(DATA_CONSULTA)
    
    print("📝 Criando arquivo Excel...")
    nome_arquivo = criar_arquivo_excel(DATA_CONSULTA, dados)
    
    if nome_arquivo:
        print(f"✅ Processo concluído com sucesso!")
    else:
        print(f"❌ Processo finalizado com erros.")